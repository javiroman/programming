from pprint import pprint

from openpyxl import load_workbook
from prettytable import PrettyTable


class IngestRequest:
    def __init__(self, filepath):
        self.wb = load_workbook(filepath)
        self.table = PrettyTable()
        self.table.field_names = ['Sheet Name', 'Rows', 'Columns']
        self.table.align['Sheet Name'] = "l"

    def print_sheet_names(self):
        print(self.wb.sheetnames)

    def print_table_summary(self):
        for i in self.wb.sheetnames:
            w = self.wb.get_sheet_by_name(i)
            self.table.add_row([i, w.max_row, w.max_column])

        print(self.table)

    def print_table_summary_real_count(self):
        for ws in self.wb.worksheets:
            for max_row, row in enumerate(ws, 1):
                if all(c.value is None for c in row):
                    break
        print(self.table)

    def print_sheet_content(self, sheet):
        ws = self.wb.get_sheet_by_name(sheet)
        pprint(list(self.__iter_rows(ws)))

    def __iter_rows(self, ws):
        for row in ws.iter_rows():
            yield [cell.value for cell in row]
